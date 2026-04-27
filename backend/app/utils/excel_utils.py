from __future__ import annotations

import io
from typing import Any, Optional

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

TEMPLATE_HEADERS = [
    "区域名称",
    "网络平面类型",
    "作用域",
    "IP地址段(CIDR)",
    "VLAN ID",
    "网关位置",
    "网关IP",
]

COLUMN_WIDTHS = [20, 16, 16, 20, 12, 30, 18]


def generate_template() -> io.BytesIO:
    """生成 Excel 导入模板。

    Returns:
        BytesIO 对象，包含格式化后的 .xlsx 模板文件。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "导入模板"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for col_idx, header in enumerate(TEMPLATE_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_idx - 1]

    ws.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def parse_excel(file_bytes: bytes) -> list[dict[str, Any]]:
    """Parse an Excel file into a list of row dicts.

    Args:
        file_bytes: Excel 文件的二进制内容。

    Returns:
        解析后的行数据列表，每行包含 region_name、plane_type_name、
        scope、ip_range、vlan_id、gateway_position、gateway_ip 等字段。
    """
    wb = load_workbook(io.BytesIO(file_bytes), read_only=True)
    ws = wb.active
    rows = []
    header_values = [str(cell.value or "").strip() for cell in ws[1]]
    has_scope_column = len(header_values) >= 3 and header_values[2] == "作用域"
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in row):
            continue
        scope = str(row[2] or "").strip() or "Global" if has_scope_column else "Global"
        ip_range_index = 3 if has_scope_column else 2
        vlan_id_index = 4 if has_scope_column else 3
        gateway_position_index = 5 if has_scope_column else 4
        gateway_ip_index = 6 if has_scope_column else 5
        rows.append(
            {
                "row_number": row_idx,
                "region_name": str(row[0] or "").strip(),
                "plane_type_name": str(row[1] or "").strip(),
                "scope": scope,
                "ip_range": str(row[ip_range_index] or "").strip(),
                "vlan_id": _parse_int(row[vlan_id_index] if len(row) > vlan_id_index else None),
                "gateway_position": (
                    str(row[gateway_position_index] or "").strip() if len(row) > gateway_position_index else ""
                )
                or None,
                "gateway_ip": (str(row[gateway_ip_index] or "").strip() if len(row) > gateway_ip_index else "") or None,
            }
        )
    wb.close()
    return rows


def _parse_int(v: Any) -> Optional[int]:
    if v is None:
        return None
    try:
        return int(v)
    except (ValueError, TypeError):
        return None


def build_export(data: list[dict[str, Any]]) -> io.BytesIO:
    """Build an Excel export workbook from Region network plane data.

    Args:
        data: 导出数据列表，每行应包含 region_name、plane_type_name、
              scope、ip_range、vlan_id、gateway_position、gateway_ip。

    Returns:
        BytesIO 对象，包含格式化后的 .xlsx 导出文件。
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "网络平面导出"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    headers = ["区域", "网络平面类型", "作用域", "IP地址段(CIDR)", "VLAN ID", "网关位置", "网关IP"]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = COLUMN_WIDTHS[col_idx - 1]

    for row_idx, item in enumerate(data, 2):
        values = [
            item.get("region_name", ""),
            item.get("plane_type_name", ""),
            item.get("scope", "Global"),
            item.get("ip_range", ""),
            item.get("vlan_id"),
            item.get("gateway_position", ""),
            item.get("gateway_ip", ""),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = thin_border

    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
