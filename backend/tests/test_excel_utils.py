"""Excel utility tests."""

import io

from openpyxl import load_workbook

from app.utils.excel_utils import generate_template, parse_excel


def test_generate_template():
    buf = generate_template()
    wb = load_workbook(buf)
    ws = wb.active

    assert ws.title == "导入模板"
    # Header row
    headers = [cell.value for cell in ws[1]]
    assert "区域名称" in headers
    assert "作用域" in headers
    assert "IP地址段(CIDR)" in headers

    wb.close()


def test_parse_empty_excel():
    """Parsing a template with no data should return empty list."""
    buf = generate_template()
    rows = parse_excel(buf.getvalue())
    assert rows == []


def test_parse_valid_excel():
    """Test parsing Excel with valid data rows."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(["区域名称", "网络平面类型", "IP地址段(CIDR)", "VLAN ID", "网关位置", "网关IP"])
    ws.append(["TestRegion", "管理平面", "10.0.0.0/24", 100, "SW01 / SW02", "10.0.0.1"])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    rows = parse_excel(buf.getvalue())
    assert len(rows) == 1
    assert rows[0]["region_name"] == "TestRegion"
    assert rows[0]["scope"] == "Global"
    assert rows[0]["ip_range"] == "10.0.0.0/24"
    assert rows[0]["vlan_id"] == 100
    assert rows[0]["gateway_position"] == "SW01 / SW02"
    assert rows[0]["gateway_ip"] == "10.0.0.1"
    wb.close()
